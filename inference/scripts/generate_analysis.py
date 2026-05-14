"""
Prediction Analysis Excel 
=================================================
Generates two Excel files (hard + soft predictions) with 4 sheets each:
  Sheet 1: Task 2.1 — Sexism Detection (binary)
  Sheet 2: Task 2.2 — Source Intention (NO/DIRECT/JUDGEMENTAL)
  Sheet 3: Task 2.3 — Sexism Categories (multilabel)
  Sheet 4: Summary  — Stats per model

Usage:
  python generate_analysis.py \
      --pred_dir    inference/predictions \
      --test_parquet data/processed/test_model_ready.parquet \
      --output_dir  inference/predictions
"""

import os, argparse
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

#Thresholds
T21   = 0.5    # Task 2.1: sexist if p21 >= T21
T23   = 0.5    # Task 2.3: category active if p_cat >= T23

# Expected Distribution based on our EDA.  Based on this we will create our thresholds.
TRAIN_FREQ = {
    "p23_ideological":     0.290,
    "p23_stereotyping":    0.225,
    "p23_objectification": 0.220,
    "p23_sexual_violence": 0.068,
    "p23_misogyny":        0.025,
}


CAT_COLS = {
    "ideological":   "p23_ideological",
    "misogyny":      "p23_misogyny",
    "objectif":      "p23_objectification",
    "sexual":        "p23_sexual_violence",
    "stereotyping":  "p23_stereotyping",
}

MODELS = ["baseline", "gated", "cross_attention", "ensemble"]
MODEL_LABELS = {
    "baseline":       "Baseline",
    "gated":          "Gated",
    "cross_attention":"CrossAttn",
    "ensemble":       "Ensemble",
}

#Format
GREEN  = PatternFill("solid", fgColor="C6EFCE")   # sexist / positive
RED    = PatternFill("solid", fgColor="FFC7CE")   # non-sexist / negative
YELLOW = PatternFill("solid", fgColor="FFEB9C")   # disagreement
BLUE_H = PatternFill("solid", fgColor="1F4E79")   # header
GREY_H = PatternFill("solid", fgColor="D9D9D9")   # subheader
WHITE  = Font(color="FFFFFF", bold=True)
BOLD   = Font(bold=True)


#Hard Prediction Helpers

# 2.1 Binary Sexist Label
def hard_21(df): 
    return (df["p21"] >= T21).astype(int)

# 2.2 Intention Hard Label
def hard_22(df, h21_series):
    """Returns NO / DIRECT / JUDGEMENTAL"""
    pred = []
    for i, row in df.iterrows():
        if h21_series.iloc[i] == 0:
            pred.append("NO")
        else:
            pred.append("DIRECT" if row["p22_direct"] >= row["p22_judgemental"] else "JUDGEMENTAL")
    return pd.Series(pred, index=df.index)


#Helper

#We calculate our thresholds based on our expected distribution
def calibrate_t23(df):
    """Compute per-category thresholds matching training frequency."""
    sexist = df[df["p21"] >= T21]
    thresholds = {}
    for col, freq in TRAIN_FREQ.items():
        thresholds[col] = float(np.percentile(sexist[col], 100 * (1 - freq)))
    return thresholds


# 2.3 Category Hard Label
def hard_23_cats(df, h21_series):
    t23_model = calibrate_t23(df)
    result = {}
    for short, col in CAT_COLS.items():
        vals = []
        for i, row in df.iterrows():
            if h21_series.iloc[i] == 0:
                vals.append(0)
            else:
                vals.append(int(row[col] >= t23_model[col]))
        result[short] = pd.Series(vals, index=df.index)
    return result


# ── Excel formatting helpers ─────────────────────────────────────────────────
def style_header_row(ws, row_num, n_cols, fill=BLUE_H, font=WHITE):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

def style_subheader_row(ws, row_num, n_cols):
    style_header_row(ws, row_num, n_cols, fill=GREY_H, font=BOLD)

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, val_len)
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

def set_text_col_width(ws, col_letter, width=50):
    ws.column_dimensions[col_letter].width = width


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_task21_sheet(ws, meta, preds_hard, preds_soft, mode="hard"):
    """Task 2.1: id | lang | text | [per model: pred] | agree"""
    models = MODELS
    # Header
    header = ["ID", "Lang", "Text"] + [MODEL_LABELS[m] for m in models] + ["Models Agree"]
    for c, h in enumerate(header, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(header))

    for i, (_, meta_row) in enumerate(meta.iterrows(), start=2):
        ws.cell(row=i, column=1, value=meta_row["id"])
        ws.cell(row=i, column=2, value=meta_row["lang"])
        ws.cell(row=i, column=3, value=str(meta_row["text"])[:200])

        preds_i = []
        for j, m in enumerate(models):
            if mode == "hard":
                val = int(preds_hard[m]["t21"].iloc[i-2])
                ws.cell(row=i, column=4+j, value=val)
                ws.cell(row=i, column=4+j).fill = GREEN if val == 1 else RED
            else:
                val = round(float(preds_soft[m]["p21"].iloc[i-2]), 4)
                ws.cell(row=i, column=4+j, value=val)
            preds_i.append(val if mode == "hard" else int(val >= T21))

        # Models agree?
        agree = len(set(preds_i)) == 1
        agree_cell = ws.cell(row=i, column=4+len(models), value=" " if agree else "✗")
        if not agree:
            agree_cell.fill = YELLOW

    auto_width(ws)
    set_text_col_width(ws, get_column_letter(3))
    ws.freeze_panes = "D2"


def build_task22_sheet(ws, meta, preds_hard, preds_soft, mode="hard"):
    """Task 2.2: id | lang | text | [per model: NO/DIRECT/JUDGEMENTAL or probs] | agree"""
    models = MODELS
    if mode == "hard":
        header = ["ID", "Lang", "Text"] + [MODEL_LABELS[m] for m in models] + ["Models Agree"]
    else:
        sub_cols = ["NO", "DIRECT", "JUDG."]
        header = ["ID", "Lang", "Text"]
        for m in models:
            header += [f"{MODEL_LABELS[m]}_{c}" for c in sub_cols]
        header += ["Models Agree (hard)"]

    for c, h in enumerate(header, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(header))

    FILL_22 = {"NO": RED, "DIRECT": GREEN, "JUDGEMENTAL": PatternFill("solid", fgColor="BDD7EE")}

    for i, (_, meta_row) in enumerate(meta.iterrows(), start=2):
        ws.cell(row=i, column=1, value=meta_row["id"])
        ws.cell(row=i, column=2, value=meta_row["lang"])
        ws.cell(row=i, column=3, value=str(meta_row["text"])[:200])

        hard_preds_i = []
        if mode == "hard":
            for j, m in enumerate(models):
                val = preds_hard[m]["t22"].iloc[i-2]
                cell = ws.cell(row=i, column=4+j, value=val)
                cell.fill = FILL_22.get(val, RED)
                hard_preds_i.append(val)
            agree = len(set(hard_preds_i)) == 1
            agree_cell = ws.cell(row=i, column=4+len(models),
                                  value=" " if agree else "✗")
            if not agree:
                agree_cell.fill = YELLOW
        else:
            col = 4
            for m in models:
                p21  = float(preds_soft[m]["p21"].iloc[i-2])
                p_no = 1 - p21
                p_d  = p21 * float(preds_soft[m]["p22_direct"].iloc[i-2])
                p_j  = p21 * float(preds_soft[m]["p22_judgemental"].iloc[i-2])
                for val in [round(p_no,4), round(p_d,4), round(p_j,4)]:
                    ws.cell(row=i, column=col, value=val)
                    col += 1
                # hard agree tracking
                hard_preds_i.append(preds_hard[m]["t22"].iloc[i-2])
            agree = len(set(hard_preds_i)) == 1
            agree_cell = ws.cell(row=i, column=col, value=" " if agree else "✗")
            if not agree:
                agree_cell.fill = YELLOW

    auto_width(ws)
    set_text_col_width(ws, get_column_letter(3))
    ws.freeze_panes = "D2"


def build_task23_sheet(ws, meta, preds_hard, preds_soft, mode="hard"):
    """Task 2.3: id | lang | text | [5 cats per model] | agree"""
    models = MODELS
    cat_short = list(CAT_COLS.keys())
    cat_labels = ["Ideol.", "Misog.", "Object.", "Sexual", "Stereot."]

    header = ["ID", "Lang", "Text"]
    for m in models:
        for cl in cat_labels:
            header.append(f"{MODEL_LABELS[m]}\n{cl}")
    header.append("Models\nAgree")

    for c, h in enumerate(header, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(header))

    for i, (_, meta_row) in enumerate(meta.iterrows(), start=2):
        ws.cell(row=i, column=1, value=meta_row["id"])
        ws.cell(row=i, column=2, value=meta_row["lang"])
        ws.cell(row=i, column=3, value=str(meta_row["text"])[:200])

        col = 4
        all_cat_preds = {cs: [] for cs in cat_short}

        for m in models:
            for cs, col_name in CAT_COLS.items():
                if mode == "hard":
                    val = int(preds_hard[m]["t23"][cs].iloc[i-2])
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.fill = GREEN if val == 1 else RED
                    all_cat_preds[cs].append(val)
                else:
                    p21 = float(preds_soft[m]["p21"].iloc[i-2])
                    val = round(p21 * float(preds_soft[m][col_name].iloc[i-2]), 4)
                    ws.cell(row=i, column=col, value=val)
                    all_cat_preds[cs].append(int(val / p21 >= T23) if p21 > 0 else 0)
                col += 1

        # Agree if all models give same prediction for ALL categories
        agree = all(len(set(v)) == 1 for v in all_cat_preds.values())
        agree_cell = ws.cell(row=i, column=col, value=" " if agree else "✗")
        if not agree:
            agree_cell.fill = YELLOW

    auto_width(ws)
    set_text_col_width(ws, get_column_letter(3))
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 35


def build_summary_sheet(ws, meta, preds_hard, preds_soft):
    """Summary stats per model."""
    ws.cell(row=1, column=1, value="EXIST 2026 — Prediction Summary")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    headers = ["Metric", "Baseline", "Gated", "CrossAttn", "Ensemble"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    metrics = []

    # Task 2.1
    metrics.append(("── Task 2.1 ──", "", "", "", ""))
    for m in MODELS:
        pass
    t21_pct = {m: f"{preds_hard[m]['t21'].mean()*100:.1f}%" for m in MODELS}
    metrics.append(("% Predicted Sexist",
                    t21_pct["baseline"], t21_pct["gated"],
                    t21_pct["cross_attention"], t21_pct["ensemble"]))

    # Agreement T21
    votes = pd.DataFrame({m: preds_hard[m]["t21"] for m in ["baseline","gated","cross_attention"]})
    agree_pct = (votes.nunique(axis=1) == 1).mean()
    metrics.append(("Models Agree (T2.1)", f"{agree_pct*100:.1f}%", "", "", ""))

    # Task 2.2
    metrics.append(("── Task 2.2 ──", "", "", "", ""))
    for label, cat in [("% NO", "NO"), ("% DIRECT", "DIRECT"), ("% JUDGEMENTAL", "JUDGEMENTAL")]:
        row = [label]
        for m in MODELS:
            pct = (preds_hard[m]["t22"] == cat).mean()
            row.append(f"{pct*100:.1f}%")
        metrics.append(tuple(row))

    # Task 2.3
    metrics.append(("── Task 2.3 ──", "", "", "", ""))
    cat_label_map = {
        "ideological": "% Ideological", "misogyny": "% Misogyny",
        "objectif": "% Objectification", "sexual": "% Sexual Violence",
        "stereotyping": "% Stereotyping"
    }
    for cs, label in cat_label_map.items():
        row = [label]
        for m in MODELS:
            pct = preds_hard[m]["t23"][cs].mean()
            row.append(f"{pct*100:.1f}%")
        metrics.append(tuple(row))

    # Write metrics
    for r, metric in enumerate(metrics, start=4):
        for c, val in enumerate(metric, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if str(val).startswith("──"):
                cell.font = Font(bold=True)
                cell.fill = GREY_H

    auto_width(ws)
    ws.column_dimensions["A"].width = 25


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pred_dir",     default="inference/predictions")
    parser.add_argument("--test_parquet", default="data/processed/test_model_ready.parquet")
    parser.add_argument("--output_dir",   default="inference/predictions")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    # Load metadata
    df_meta = pd.read_parquet(args.test_parquet)[["id", "lang", "text"]]

    # Load predictions
    preds_soft = {}
    for m in MODELS:
        path = os.path.join(args.pred_dir, f"{m}_raw.parquet")
        preds_soft[m] = pd.read_parquet(path).set_index("id").loc[df_meta["id"]].reset_index()

    # Compute hard predictions
    preds_hard = {}
    for m in MODELS:
        df = preds_soft[m]
        h21 = hard_21(df)
        preds_hard[m] = {
            "t21": h21,
            "t22": hard_22(df, h21),
            "t23": hard_23_cats(df, h21),
        }

    print(f"  Memes: {len(df_meta)}")
    for m in MODELS:
        pct = preds_hard[m]["t21"].mean()
        print(f"  {MODEL_LABELS[m]:12s} → {pct*100:.1f}% predicted sexist")

    # Generate both Excel files
    for mode in ["hard", "soft"]:
        out_path = os.path.join(args.output_dir, f"predictions_{mode}.xlsx")
        print(f"\nGenerating {out_path}...")

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            # Write placeholder sheets (openpyxl will manage them)
            pd.DataFrame(["placeholder"]).to_excel(writer, sheet_name="Task 2.1", index=False)
            pd.DataFrame(["placeholder"]).to_excel(writer, sheet_name="Task 2.2", index=False)
            pd.DataFrame(["placeholder"]).to_excel(writer, sheet_name="Task 2.3", index=False)
            pd.DataFrame(["placeholder"]).to_excel(writer, sheet_name="Summary",  index=False)

        # Re-open with openpyxl for proper formatting
        wb = load_workbook(out_path)
        sheets = {"Task 2.1": wb["Task 2.1"], "Task 2.2": wb["Task 2.2"],
                  "Task 2.3": wb["Task 2.3"], "Summary": wb["Summary"]}

        # Clear placeholder content
        for ws in sheets.values():
            ws.delete_rows(1, ws.max_row)

        print("  Building Task 2.1 sheet...")
        build_task21_sheet(sheets["Task 2.1"], df_meta, preds_hard, preds_soft, mode)

        print("  Building Task 2.2 sheet...")
        build_task22_sheet(sheets["Task 2.2"], df_meta, preds_hard, preds_soft, mode)

        print("  Building Task 2.3 sheet...")
        build_task23_sheet(sheets["Task 2.3"], df_meta, preds_hard, preds_soft, mode)

        print("  Building Summary sheet...")
        build_summary_sheet(sheets["Summary"], df_meta, preds_hard, preds_soft)

        wb.save(out_path)
        print(f"    Saved: {out_path}")

    print("\n  Done. Check inference/predictions/predictions_hard.xlsx and predictions_soft.xlsx")


if __name__ == "__main__":
    main()
